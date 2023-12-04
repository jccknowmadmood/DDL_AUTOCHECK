import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def limpiar_cadena_create(cadena):
    # Quitar todas las comillas dobles
    cadena = cadena.replace('"', '')

    # Quitar las subcadenas " BYTE"
    cadena = cadena.replace(' BYTE', '')

    # Quitar las subcadenas "NOT NULL"
    cadena = cadena.replace('NOT NULL', '')

    # Sustituir todas las subcadenas ", " por ","
    cadena = cadena.replace(', ', ',')

    # Sustituir todas las subcadenas " ," por ","
    cadena = cadena.replace(' ,', ',')

    cadena = cadena.replace('TIMESTAMP (6)','TIMESTAMP ')
    
    cadena = cadena.replace('  ',' ')

    # Eliminar todo lo que haya en la cadena a partir de "SEGMENT CREATION"
    indice_segment_creation = cadena.find('SEGMENT CREATION')
    if indice_segment_creation != -1:
        cadena = cadena[:indice_segment_creation]

    cadena=cadena.strip()    

    return cadena


# Abre el archivo SQL en modo lectura con codificación UTF-8
with open('DDL_SI.sql', 'r', encoding='utf-8') as archivo:
    contenido = archivo.read()

# Divide el contenido en declaraciones usando punto y coma como separador
declaraciones = contenido.split(';')

# Expresiones regulares para extraer nombre de tabla y atributos
declaracion_tabla = r'CREATE TABLE\s+(.*?)\s*\('
declaracion_alter = r'ALTER TABLE\s+(.*?)\s'
declaracion_comment = r'COMMENT ON\s+(.*?)\s'

tablas_SI = []
alter_tables_SI = []
comment_tables_SI = []
errores_SI = []

for declaracion in declaraciones:
    nombre_tabla_match = re.search(declaracion_tabla, declaracion)
    alter_match = re.search(declaracion_alter, declaracion)
    comment_match = re.search(declaracion_comment, declaracion)

    if nombre_tabla_match:
        nombre_tabla = nombre_tabla_match.group(1)

        indice_parentesis = declaracion.find('(')
        bloque_atributos = (declaracion[indice_parentesis:])[2:-1]
        bloque_atributos = bloque_atributos.splitlines()
        attribute_list = []        

        for linea in bloque_atributos:
            linea = re.sub(r'\s{2,}', ' ', linea)
            linea = linea.strip()
            if linea =="," or linea=="":
                continue
            if "ERROR" in linea:
                errores_SI.append({"Tabla": nombre_tabla.upper().strip(), "Error": linea})
                continue
            nombre_atributo, tipo_atributo = linea.split(' ', 1)
            #print("nombre atb: "+ nombre_atributo + " / tipo atributo: " + tipo_atributo)
            attribute_list.append({'nombre_atributo': nombre_atributo.upper().strip(), 'tipo_atributo': tipo_atributo.upper().strip()})
        declaracion = re.sub(r'\n', ' ', declaracion)        
        declaracion = re.sub(r'(?<=[^\' ]) +| +(?=[^\' ])', ' ', declaracion)
        tablas_SI.append({'nombre_tabla': nombre_tabla.upper().strip(), 'atributos': attribute_list, 'ddl': declaracion.upper().strip()})   

    elif alter_match:
        nombre_tabla = declaracion.split(' ')[2]
        nombre_tabla = re.sub(r'\n', '', nombre_tabla)
        declaracion = re.sub(r'\n', ' ', declaracion)        
        declaracion = re.sub(r'(?<=[^\' ]) +| +(?=[^\' ])', ' ', declaracion) # sustituye grupos de espacios por uno solo, a no ser que estén entre comillas
        alter_tables_SI.append((nombre_tabla.upper().strip(), declaracion.upper().strip()))    

    elif comment_match:
        nombre_tabla = declaracion.split(' ')[3]
        if '.' in nombre_tabla:
            nombre_tabla = nombre_tabla.split('.')[0]
        nombre_tabla = re.sub(r'\n', '', nombre_tabla)
        declaracion = re.sub(r'\n', ' ', declaracion)        
        declaracion = re.sub(r'(?<=[^\' ]) +| +(?=[^\' ])', ' ', declaracion) # sustituye grupos de espacios por uno solo, a no ser que estén entre comillas
        comment_tables_SI.append((nombre_tabla.upper().strip(), declaracion.upper().strip()))   


with open('ddl_db_desa.sql', 'r') as archivo: #ddl_db_desa_PRUEBA.sql
    sql_script = archivo.read()

table_definition_pattern = r'CREATE TABLE[^;]*'
declaracion_alter = r'ALTER TABLE "OEVDSSII"\."([^"]+)"(.*?)"'
declaracion_comment = r'COMMENT ON\s+(.*?)\s'

declaraciones = re.split(';', sql_script)

tablas_DB_desa = []
alter_tables_DB_desa = []
comment_tables_DB_desa = []
errores_DB_desa = []

for declaracion in declaraciones:

    create_tabla_match = re.search(table_definition_pattern, declaracion)
    alter_tabla_match = re.search(declaracion_alter, declaracion)
    comment_tabla_match = re.search(declaracion_comment, declaracion)


    if create_tabla_match:
        table_name_match = re.search(r'CREATE TABLE "OEVDSSII"\."([^"]+)"', declaracion)        
        if table_name_match:
            table_name = table_name_match.group(1)
            attributes_match = re.search(rf'{table_name}[^;]*\) SEGMENT', declaracion)
            if attributes_match:
                attributes_text = attributes_match.group(0)
                attribute_lines = attributes_text.split('\n')
                attribute_list = []
                for line in attribute_lines:
                    attribute_match = re.search(r'"([^"]+)"\s+([^,]+),', line)
                    if attribute_match:
                        attribute_name = attribute_match.group(1)
                        attribute_type = attribute_match.group(2)
                        attribute_list.append({'nombre_atributo': attribute_name.upper().strip(), 'tipo_atributo': attribute_type.upper().strip()})

                declaracion = declaracion.split('\n')
                filtered_lines = [line for line in declaracion if not line.strip().startswith('-')]

                # Unir las líneas restantes en una sola línea
                result_string = ' '.join(filtered_lines)
                result_string = re.sub(r'\s{2,}', ' ', result_string)
                # Eliminar cualquier ocurrencia de "OEVDSSII".
                result_string = result_string.replace('"OEVDSSII".', '')        
                
                result_string = re.sub(r'\t', ' ', result_string)         
                result_string = re.sub(r'(?<=[^\' ]) +| +(?=[^\' ])', ' ', result_string)        
                tablas_DB_desa.append({'nombre_tabla': table_name.upper().strip(), 'atributos': attribute_list, 'ddl' : result_string.upper().strip()})

    elif alter_tabla_match:
        alter_statement = declaracion.split('\n')
        filtered_lines = [line for line in alter_statement if not line.strip().startswith('-')]

        # Unir las líneas restantes en una sola línea
        result_string = ' '.join(filtered_lines)
        result_string = re.sub(r'\s{2,}', ' ', result_string)
        # Eliminar cualquier ocurrencia de "OEVDSSII".
        result_string = result_string.replace('"OEVDSSII".', '')

        match = re.search(r'"(.*?)"', result_string)
        if match:
            table_name = match.group(1)
        alter_tables_DB_desa.append((table_name, result_string.upper().strip()))

    elif comment_tabla_match:
        comment_statement = declaracion.split('\n')
        filtered_lines = [line for line in comment_statement if not line.strip().startswith('-')]

        result_string = declaracion.replace('"OEVDSSII".', '')
        result_string = result_string.replace('\n', ' ')

        # Dividir la cadena en dos partes en la subcadena "IS"
        parts = result_string.split("IS", 1)
        # Eliminar todas las comillas de la primera subcadena
        first_part = parts[0].replace('"', '')
        # Volver a unir las partes
        result_string = first_part + "IS" + parts[1]
        # Sustituir cualquier cadena de más de un espacio por un solo espacio
        result_string = re.sub(r'\s{2,}', ' ', result_string)
                
        # Crear la variable table_name
        if result_string.startswith("COMMENT ON COLUMN"):
            table_name = result_string.split(' ', 3)[3].split('.')[0]
        elif result_string.startswith("COMMENT ON TABLE"):
            table_name = result_string.split(' ', 4)[3]

        comment_tables_DB_desa.append((table_name.upper().strip(), result_string.upper().strip()))     


# Extraer los table_name y DDL de ambas listas
table_names_SI = {tabla["nombre_tabla"]: tabla["ddl"] for tabla in tablas_SI}
table_names_DB_desa = {tabla["nombre_tabla"]: tabla["ddl"] for tabla in tablas_DB_desa}

# Encontrar elementos presentes solo en tablas_SI
elementos_solo_en_SI = set(table_names_SI.keys()) - set(table_names_DB_desa.keys())

# Encontrar elementos presentes solo en tablas_DB_desa
elementos_solo_en_DB_desa = set(table_names_DB_desa.keys()) - set(table_names_SI.keys())

# Encontrar los elementos comunes (claves en ambos diccionarios)
elementos_comunes = set(table_names_SI.keys()) & set(table_names_DB_desa.keys())

# Obtener el campo 'DDL' correspondiente para los elementos solo en tablas_SI
sentencias_create_para_DB_desa = {elemento: table_names_SI[elemento] for elemento in elementos_solo_en_SI}
# Obtener el campo 'DDL' correspondiente para los elementos solo en tablas_DB_desa
sentencias_create_para_SI = {elemento: table_names_DB_desa[elemento] for elemento in elementos_solo_en_DB_desa}

# Convertir los conjuntos a listas para su uso en DataFrames
elementos_comunes_lista = list(elementos_comunes)
elementos_solo_en_SI_lista = list(elementos_solo_en_SI)
elementos_solo_en_DB_desa_lista = list(elementos_solo_en_DB_desa)
sentencias_create_para_DB_desa_lista = list(sentencias_create_para_DB_desa.items())
sentencias_create_para_SI_lista = list(sentencias_create_para_SI.items())

# Crear DataFrames para cada conjunto o lista
df_elementos_comunes = pd.DataFrame({"Elementos Comunes": elementos_comunes_lista})
df_solo_tablas_SI = pd.DataFrame({"Elementos Solo en tablas_SI": elementos_solo_en_SI_lista})
df_solo_tablas_DB_desa = pd.DataFrame({"Elementos Solo en tablas_DB_desa": elementos_solo_en_DB_desa_lista})
df_ddl_create_para_DB_desa = pd.DataFrame(sentencias_create_para_DB_desa_lista, columns=["Tabla", "DDL"])
df_ddl_create_para_SI = pd.DataFrame(sentencias_create_para_SI_lista, columns=["Tabla", "DDL"])


nombres_tablas_DB_desa = df_ddl_create_para_DB_desa["Tabla"].tolist()
nombres_tablas_SI = df_ddl_create_para_SI["Tabla"].tolist()

sentencias_alter_relacionadas = [fila for fila in alter_tables_SI if fila[0] in nombres_tablas_DB_desa]
sentencias_alter_para_DB_desa = [fila[1] for fila in sentencias_alter_relacionadas]
df_alter_para_DB_desa = pd.DataFrame(sentencias_alter_para_DB_desa, columns=["Sentencias ALTER para DB_desa"])

sentencias_alter_relacionadas = [fila for fila in alter_tables_DB_desa if fila[0] in nombres_tablas_SI]
sentencias_alter_para_SI = [fila[1] for fila in sentencias_alter_relacionadas]
df_alter_para_SI = pd.DataFrame(sentencias_alter_para_SI, columns=["Sentencias ALTER para SI"])

sentencias_comment_relacionadas = [fila for fila in comment_tables_SI if fila[0] in nombres_tablas_DB_desa]
sentencias_comment_para_DB_desa = [fila[1] for fila in sentencias_comment_relacionadas]
df_comment_para_DB_desa = pd.DataFrame(sentencias_comment_para_DB_desa, columns=["Sentencias COMMENT para DB_desa"])

sentencias_comment_relacionadas = [fila for fila in comment_tables_DB_desa if fila[0] in nombres_tablas_SI]
sentencias_comment_para_SI = [fila[1] for fila in sentencias_comment_relacionadas]
df_comment_para_SI = pd.DataFrame(sentencias_comment_para_SI, columns=["Sentencias COMMENT para SI"])





#TODO Ver si los alter table son iguales
#TODO ver si los comentarios son iguales

estados_tablas = []

for elemento in elementos_comunes:

    ddl_si = limpiar_cadena_create(table_names_SI[elemento])
    ddl_db_desa = limpiar_cadena_create(table_names_DB_desa[elemento])

    if ddl_si == ddl_db_desa:
        estados_tablas.append('OK')
    else:
        estados_tablas.append('KO')


df_elementos_comunes['Estado'] = estados_tablas      

fecha_actual = datetime.now()
fecha_str = fecha_actual.strftime("%d%m%Y")
excel_writer = pd.ExcelWriter(f'INFORME_BBDD_SI-DB_DESA_{fecha_str}.xlsx', engine='xlsxwriter')

# Escribir los DataFrames en hojas separadas
df_elementos_comunes.to_excel(excel_writer, sheet_name='Elementos comunes', index=False)
df_solo_tablas_SI.to_excel(excel_writer, sheet_name='Solo tablas_SI', index=False)
df_solo_tablas_DB_desa.to_excel(excel_writer, sheet_name='Solo tablas_DB_desa', index=False)
df_ddl_create_para_SI.to_excel(excel_writer, sheet_name='DDL TBLs no en SI', index=False) 
df_ddl_create_para_DB_desa.to_excel(excel_writer, sheet_name='DDL TBLs no en DB_desa', index=False)
df_alter_para_SI.to_excel(excel_writer, sheet_name='ALTER inexistentes en SI', index=False) #Las que hay que hacer obligatoriamente porque no existían previamente
df_alter_para_DB_desa.to_excel(excel_writer, sheet_name='ALTER inexistentes en DB_desa', index=False) #Las que hay que hacer obligatoriamente porque no existían previamente
df_comment_para_SI.to_excel(excel_writer, sheet_name='COMMENT inexistentes en SI', index=False) #Los que hay que hacer obligatoriamente porque no existían previamente
df_comment_para_DB_desa.to_excel(excel_writer, sheet_name='COMMENT inexistentes en DB_desa', index=False) #Los que hay que hacer obligatoriamente porque no existían previamente

# Guardar el archivo Excel
excel_writer._save()


colores = {'OK': 'b6d7a8', 'KO': 'fd5b5b'}

# Cargar el archivo Excel para manipular el formato de las celdas
archivo_excel = load_workbook(f'INFORME_BBDD_SI-DB_DESA_{fecha_str}.xlsx')

# Acceder a la hoja de 'Elementos comunes'
hoja_elementos_comunes = archivo_excel['Elementos comunes']

# Obtener las celdas de la columna 'Estado'
celdas_estado = hoja_elementos_comunes['B']

# Iterar sobre las celdas y aplicar el formato
for celda in celdas_estado:
    estado = celda.value
    if estado in colores:
        fill = PatternFill(start_color=colores[estado], end_color=colores[estado], fill_type='solid')
        celda.fill = fill

hoja_elementos_comunes.auto_filter.ref = hoja_elementos_comunes.dimensions

# Guardar el archivo Excel actualizado
archivo_excel.save(f'INFORME_BBDD_SI-DB_DESA_{fecha_str}.xlsx')

print("Proceso Finalizado")
